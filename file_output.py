"""
File output module for CSV and Excel generation.
Creates portfolio snapshots and historical tracking.
"""

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils import (
    Holding, round_currency, format_currency, format_percent, get_date_str
)
from portfolio import PortfolioSummary, HoldingAnalysis
from recommendations import RecommendationReport, ActionType


logger = logging.getLogger('portfolio_automation.file_output')


# Style constants for Excel
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALERT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def ensure_directory(path: Path) -> None:
    """Ensure parent directory exists."""
    path.parent.mkdir(parents=True, exist_ok=True)


def write_csv_snapshot(
    filepath: str,
    holdings: list[Holding],
    analyses: list[HoldingAnalysis],
    summary: PortfolioSummary,
    cash_available: float
) -> bool:
    """
    Write portfolio snapshot to CSV file.
    Returns True if successful.
    """
    path = Path(filepath)
    ensure_directory(path)
    
    try:
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([
                'Symbol', 'Shares', 'Price', 'Market_Value',
                'Target_Weight', 'Actual_Weight', 'Drift',
                'Asset_Class', 'Is_Leveraged', 'Leverage_Factor',
                'Status', 'Timestamp'
            ])
            
            # Holdings data
            for holding, analysis in zip(holdings, analyses):
                status = 'BREACH' if analysis.is_breached else analysis.drift_direction
                
                writer.writerow([
                    holding.symbol,
                    round(holding.shares, 4),
                    round(holding.current_price or 0, 2),
                    round(holding.market_value or 0, 2),
                    round(holding.target_weight, 4),
                    round(analysis.actual_weight, 4),
                    round(analysis.drift, 4),
                    holding.asset_class,
                    holding.is_leveraged,
                    holding.leverage_factor,
                    status,
                    summary.timestamp
                ])
            
            # Cash row
            writer.writerow([
                'CASH', cash_available, 1.0, cash_available,
                '', summary.cash_weight, '',
                'cash', False, 1.0,
                'N/A', summary.timestamp
            ])
            
            # Summary section
            writer.writerow([])
            writer.writerow(['SUMMARY'])
            writer.writerow(['Total Holdings Value', summary.total_holdings_value])
            writer.writerow(['Cash Available', summary.cash_value])
            writer.writerow(['Total Portfolio', summary.total_portfolio_value])
            writer.writerow(['401(k) Balance', summary.retirement_401k_value])
            writer.writerow(['Total Net Worth', summary.total_net_worth])
            writer.writerow(['Max Drift', f"{summary.max_drift:.4f}"])
            writer.writerow(['Max Drift Symbol', summary.max_drift_symbol])
            writer.writerow(['Rebalance Needed', summary.has_breach])
        
        logger.info(f"CSV snapshot written: {filepath}")
        return True
        
    except IOError as e:
        logger.error(f"Failed to write CSV: {e}")
        return False


def create_excel_workbook(
    filepath: str,
    holdings: list[Holding],
    analyses: list[HoldingAnalysis],
    summary: PortfolioSummary,
    recommendations: RecommendationReport,
    cash_available: float,
    append_history: bool = True
) -> bool:
    """
    Create or update Excel workbook with portfolio data.
    Includes Holdings, Summary, and History sheets.
    Returns True if successful.
    """
    path = Path(filepath)
    ensure_directory(path)
    
    try:
        # Load existing or create new workbook
        if path.exists() and append_history:
            wb = load_workbook(path)
        else:
            wb = Workbook()
        
        # Create/update Holdings sheet
        _write_holdings_sheet(wb, holdings, analyses, summary)
        
        # Create/update Summary sheet
        _write_summary_sheet(wb, summary, recommendations, cash_available)
        
        # Append to History sheet
        if append_history:
            _append_history_sheet(wb, holdings, summary)
        
        # Remove default sheet if empty
        if 'Sheet' in wb.sheetnames and wb['Sheet'].max_row == 1:
            del wb['Sheet']
        
        wb.save(path)
        logger.info(f"Excel workbook saved: {filepath}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to create Excel workbook: {e}")
        return False


def _write_holdings_sheet(
    wb: Workbook,
    holdings: list[Holding],
    analyses: list[HoldingAnalysis],
    summary: PortfolioSummary
) -> None:
    """Write or update Holdings sheet."""
    sheet_name = 'Holdings'
    
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name, 0)
    
    # Headers
    headers = [
        'Symbol', 'Shares', 'Price', 'Market Value',
        'Target %', 'Actual %', 'Drift', 'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    # Data rows
    for row_idx, (holding, analysis) in enumerate(zip(holdings, analyses), 2):
        status = 'BREACH' if analysis.is_breached else analysis.drift_direction.upper()
        
        row_data = [
            holding.symbol,
            holding.shares,
            holding.current_price or 0,
            holding.market_value or 0,
            holding.target_weight,
            analysis.actual_weight,
            analysis.drift,
            status
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            
            # Format specific columns
            if col == 3:  # Price
                cell.number_format = '$#,##0.00'
            elif col == 4:  # Market Value
                cell.number_format = '$#,##0.00'
            elif col in [5, 6, 7]:  # Percentages
                cell.number_format = '0.00%'
            
            # Highlight breached rows
            if analysis.is_breached:
                cell.fill = ALERT_FILL
    
    # Add cash row
    cash_row = len(holdings) + 2
    cash_data = ['CASH', summary.cash_value, 1.0, summary.cash_value, '', summary.cash_weight, '', '']
    
    for col, value in enumerate(cash_data, 1):
        cell = ws.cell(row=cash_row, column=col, value=value)
        cell.border = THIN_BORDER
        if col in [2, 4]:
            cell.number_format = '$#,##0.00'
        elif col == 6:
            cell.number_format = '0.00%'
    
    # Auto-adjust column widths
    for col_idx, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


def _write_summary_sheet(
    wb: Workbook,
    summary: PortfolioSummary,
    recommendations: RecommendationReport,
    cash_available: float
) -> None:
    """Write or update Summary sheet."""
    sheet_name = 'Summary'
    
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name, 1)
    
    # Title
    ws['A1'] = 'Portfolio Summary'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    # Timestamp
    ws['A2'] = 'Last Updated:'
    ws['B2'] = summary.timestamp
    
    # Values section
    ws['A4'] = 'VALUES'
    ws['A4'].font = Font(bold=True)
    
    value_data = [
        ('Holdings Value', summary.total_holdings_value),
        ('Cash Available', summary.cash_value),
        ('Total Portfolio', summary.total_portfolio_value),
        ('401(k) Balance', summary.retirement_401k_value),
        ('Total Net Worth', summary.total_net_worth),
    ]
    
    for idx, (label, value) in enumerate(value_data, 5):
        ws[f'A{idx}'] = label
        ws[f'B{idx}'] = value
        ws[f'B{idx}'].number_format = '$#,##0.00'
    
    # Drift section
    ws['A11'] = 'DRIFT ANALYSIS'
    ws['A11'].font = Font(bold=True)
    
    ws['A12'] = 'Max Drift'
    ws['B12'] = summary.max_drift
    ws['B12'].number_format = '0.00%'
    
    ws['A13'] = 'Drift Symbol'
    ws['B13'] = summary.max_drift_symbol
    
    ws['A14'] = 'Threshold'
    ws['B14'] = summary.breach_threshold
    ws['B14'].number_format = '0.00%'
    
    ws['A15'] = 'Rebalance Needed'
    ws['B15'] = 'YES' if summary.has_breach else 'NO'
    if summary.has_breach:
        ws['B15'].fill = ALERT_FILL
    else:
        ws['B15'].fill = GOOD_FILL
    
    # Recommendations section
    ws['A17'] = 'RECOMMENDATIONS'
    ws['A17'].font = Font(bold=True)
    
    ws['A18'] = recommendations.summary_message
    ws.merge_cells('A18:C18')
    
    row = 19
    for rec in recommendations.recommendations:
        if rec.action_type in [ActionType.BUY, ActionType.SELL, ActionType.REBALANCE_ALERT]:
            ws[f'A{row}'] = rec.action_type.value
            ws[f'B{row}'] = rec.symbol
            ws[f'C{row}'] = rec.reason[:50] + '...' if len(rec.reason) > 50 else rec.reason
            
            if rec.action_type == ActionType.REBALANCE_ALERT:
                ws[f'A{row}'].fill = ALERT_FILL
            
            row += 1
    
    # Notes section
    if recommendations.notes:
        row += 1
        ws[f'A{row}'] = 'NOTES'
        ws[f'A{row}'].font = Font(bold=True)
        
        for note in recommendations.notes:
            row += 1
            ws[f'A{row}'] = f'• {note}'
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50


def _append_history_sheet(
    wb: Workbook,
    holdings: list[Holding],
    summary: PortfolioSummary
) -> None:
    """Append current snapshot to History sheet."""
    sheet_name = 'History'
    
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        
        # Create headers
        headers = ['Date', 'Total Value', 'Cash', '401k', 'Net Worth', 'Max Drift', 'Rebalance']
        for holding in holdings:
            headers.append(f'{holding.symbol}_value')
            headers.append(f'{holding.symbol}_weight')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
    else:
        ws = wb[sheet_name]
    
    # Find next row
    next_row = ws.max_row + 1
    
    # Build data row
    row_data = [
        get_date_str(),
        summary.total_portfolio_value,
        summary.cash_value,
        summary.retirement_401k_value,
        summary.total_net_worth,
        summary.max_drift,
        'YES' if summary.has_breach else 'NO'
    ]
    
    for holding in holdings:
        row_data.append(holding.market_value or 0)
        row_data.append(holding.actual_weight or 0)
    
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col, value=value)
        
        # Format numbers
        if col in [2, 3, 4, 5]:
            cell.number_format = '$#,##0.00'
        elif col == 6:
            cell.number_format = '0.00%'
        elif col > 7 and (col - 7) % 2 == 1:  # Value columns
            cell.number_format = '$#,##0.00'
        elif col > 7 and (col - 7) % 2 == 0:  # Weight columns
            cell.number_format = '0.00%'


def write_contribution_plan_csv(
    filepath: str,
    allocations: list,  # List[ContributionAllocation]
    drawdown_regime: str = 'normal',
) -> bool:
    """
    Write the contribution plan to a CSV file.

    Columns: Symbol, AssetClass, CurrentWeight, TargetWeight, Drift,
             RecommendedContributionDollars, Reason

    Returns True if successful.
    """
    path = Path(filepath)
    ensure_directory(path)

    try:
        if not allocations:
            logger.info("No contribution allocations to write")
            return False

        rows = [a.to_dict() for a in allocations]
        # Append metadata row
        rows.append({k: '' for k in rows[0]})
        rows.append({'Symbol': '# DrawdownRegime', 'Reason': drawdown_regime,
                     **{k: '' for k in rows[0] if k not in ('Symbol', 'Reason')}})

        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)

        logger.info(f"Contribution plan written: {filepath}")
        return True

    except IOError as e:
        logger.error(f"Failed to write contribution plan: {e}")
        return False


def write_compounding_dashboard_txt(
    filepath: str,
    dashboard_text: str,
) -> bool:
    """Write the formatted compounding dashboard text to a file."""
    path = Path(filepath)
    ensure_directory(path)
    try:
        with open(path, 'w', encoding='utf-8') as f:
            f.write(dashboard_text)
        logger.info(f"Compounding dashboard written: {filepath}")
        return True
    except IOError as e:
        logger.error(f"Failed to write compounding dashboard: {e}")
        return False


def export_recommendations_csv(
    filepath: str,
    recommendations: RecommendationReport,
    summary: PortfolioSummary
) -> bool:
    """Export recommendations to CSV file."""
    path = Path(filepath)
    ensure_directory(path)
    
    try:
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            writer.writerow([
                'Action', 'Symbol', 'Shares', 'Amount',
                'Reason', 'Priority', 'Urgent', 'Timestamp'
            ])
            
            for rec in recommendations.recommendations:
                # Replace ± with +/- for better compatibility
                reason = rec.reason.replace('±', '+/-')
                writer.writerow([
                    rec.action_type.value,
                    rec.symbol,
                    rec.shares or '',
                    rec.amount or '',
                    reason,
                    rec.priority,
                    rec.is_urgent,
                    summary.timestamp
                ])
        
        logger.info(f"Recommendations CSV exported: {filepath}")
        return True
        
    except IOError as e:
        logger.error(f"Failed to export recommendations: {e}")
        return False